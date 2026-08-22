from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from stock_valuation.data.types import NormalizedFinancialFact


ASML_2025_US_GAAP_XLSX = (
    "https://ourbrand.asml.com/m/6cd86f972a9dfd24/original/"
    "2025-US-GAAP-Financial-Statements.xlsx"
)


class ASMLPrimarySourceError(RuntimeError):
    """Raised when an official ASML primary-source workbook cannot be loaded or inspected."""


TARGET_PATTERNS: dict[str, tuple[str, ...]] = {
    "cash_and_equivalents": ("cash and cash equivalents",),
    "short_term_investments": ("short-term investments", "short term investments"),
    "accounts_receivable": ("accounts receivable, net", "accounts receivable"),
    "inventory": ("inventories, net", "inventories"),
    "ppe_net": (
        "property, plant and equipment, net",
        "property plant and equipment, net",
    ),
    "short_term_debt": (
        "short-term borrowings",
        "short term borrowings",
        "current portion of long-term debt",
        "current portion of long term debt",
    ),
    "operating_cash_flow": (
        "net cash provided by operating activities",
        "net cash from operating activities",
    ),
    "capital_expenditures": (
        "purchases of property, plant and equipment",
        "purchase of property, plant and equipment",
        "purchases of property plant and equipment",
    ),
    "intangible_purchases": (
        "purchases of intangible assets",
        "purchase of intangible assets",
    ),
    "dividends_paid": ("dividend paid", "dividends paid"),
}


PRIMARY_IMPORT_ROWS: tuple[tuple[str, str, tuple[str, ...], str], ...] = (
    ("balance_sheet", "cash_and_equivalents", ("cash and cash equivalents",), "Balance Sheets"),
    ("balance_sheet", "short_term_investments", ("short-term investments",), "Balance Sheets"),
    ("balance_sheet", "accounts_receivable", ("accounts receivable, net",), "Balance Sheets"),
    ("balance_sheet", "inventory", ("inventories, net",), "Balance Sheets"),
    (
        "balance_sheet",
        "ppe_net",
        ("property, plant and equipment, net",),
        "Balance Sheets",
    ),
    (
        "balance_sheet",
        "short_term_debt",
        ("short-term borrowings and current portion of long-term debt",),
        "Balance Sheets",
    ),
    (
        "cash_flow",
        "operating_cash_flow",
        ("net cash provided by operating activities",),
        "Cash Flow",
    ),
    (
        "cash_flow",
        "capital_expenditures",
        ("purchase of property, plant and equipment", "purchases of property, plant and equipment"),
        "Cash Flow",
    ),
    (
        "cash_flow",
        "intangible_purchases",
        ("purchase of intangible assets", "purchases of intangible assets"),
        "Cash Flow",
    ),
    ("cash_flow", "dividends_paid", ("dividend paid",), "Cash Flow"),
)

OUTFLOW_METRICS = {"capital_expenditures", "intangible_purchases", "dividends_paid"}


def download_2025_us_gaap_workbook(*, timeout: int = 30) -> bytes:
    """Download ASML's official 2025 US-GAAP financial-statements workbook.

    This request goes directly to ASML and does not use an Alpha Vantage API quota.
    """
    try:
        response = requests.get(
            ASML_2025_US_GAAP_XLSX,
            timeout=timeout,
            headers={"User-Agent": "stock-valuation-tool/0.1"},
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise ASMLPrimarySourceError(
            "Die offizielle ASML-US-GAAP-Excel-Datei konnte nicht geladen werden."
        ) from exc

    content = response.content
    if not content.startswith(b"PK"):
        raise ASMLPrimarySourceError(
            "ASML lieferte keine erkennbare XLSX-Datei zurück."
        )
    return content


def _row_text(values: tuple[Any, ...]) -> str:
    return " | ".join(str(value).strip() for value in values if value not in (None, ""))


def _row_cells(values: tuple[Any, ...], row_number: int) -> str:
    parts: list[str] = []
    for index, value in enumerate(values, start=1):
        if value in (None, ""):
            continue
        coordinate = f"{get_column_letter(index)}{row_number}"
        parts.append(f"{coordinate}={value}")
    return " | ".join(parts)


def _decimal(value: Any) -> Decimal | None:
    if value in (None, "", "None", "null") or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _find_exact_row(sheet, patterns: tuple[str, ...]) -> tuple[int, tuple[Any, ...]] | None:
    """Find one financial-statement row by its label in column A.

    The primary-source importer is intentionally stricter than the diagnostic scanner: only
    column-A labels are accepted, so a similarly named cash-flow movement cannot be mistaken
    for a balance-sheet closing balance.
    """
    wanted = tuple(pattern.strip().lower() for pattern in patterns)
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = tuple(row)
        label = str(values[0] or "").strip().lower() if values else ""
        if any(label == pattern for pattern in wanted):
            return row_number, values
    return None


def _last_two_numeric(values: tuple[Any, ...]) -> tuple[Decimal, Decimal] | None:
    numeric: list[Decimal] = []
    for value in values[1:]:
        parsed = _decimal(value)
        if parsed is not None:
            numeric.append(parsed)
    if len(numeric) < 2:
        return None
    return numeric[-2], numeric[-1]


def parse_primary_source_facts(
    content: bytes,
    *,
    retrieved_at: datetime | None = None,
) -> list[NormalizedFinancialFact]:
    """Parse validated 2024/2025 rows from ASML's official 2025 US-GAAP workbook.

    The live workbook has two balance-sheet comparison columns (2024/2025) and at least three
    cash-flow periods (2023/2024/2025). For the current validation gate we deliberately import
    only the two right-most numeric values, i.e. 2024 and 2025. Raw workbook values are in
    EUR millions and are preserved in `provider_value`; normalized `value` is stored in EUR.
    """
    retrieved_at = retrieved_at or datetime.now(timezone.utc)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ASMLPrimarySourceError("Die ASML-XLSX-Datei konnte nicht gelesen werden.") from exc

    facts: list[NormalizedFinancialFact] = []
    try:
        for statement, metric, patterns, sheet_name in PRIMARY_IMPORT_ROWS:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            match = _find_exact_row(sheet, patterns)
            if match is None:
                continue
            row_number, values = match
            pair = _last_two_numeric(values)
            if pair is None:
                continue

            for year, raw_millions in zip((2024, 2025), pair, strict=True):
                normalized_millions = abs(raw_millions) if metric in OUTFLOW_METRICS else raw_millions
                facts.append(
                    NormalizedFinancialFact(
                        statement=statement,
                        metric=metric,
                        period_end=date(year, 12, 31),
                        period_type="FY",
                        value=normalized_millions * Decimal("1000000"),
                        provider_value=raw_millions,
                        currency="EUR",
                        unit="currency",
                        provider="asml_primary",
                        provider_field=f"{sheet_name}!A{row_number}",
                        filing_date=None,
                        retrieved_at=retrieved_at,
                        is_cross_check_only=False,
                        note=(
                            f"Official ASML 2025 US-GAAP workbook; row label: {values[0]}. "
                            "provider_value is EUR millions; value is normalized to EUR."
                        ),
                    )
                )
    finally:
        workbook.close()
    return facts


def scan_financial_statement_workbook(content: bytes) -> list[dict[str, Any]]:
    """Find relevant official ASML rows without assuming a fixed workbook layout."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ASMLPrimarySourceError("Die ASML-XLSX-Datei konnte nicht gelesen werden.") from exc

    matches: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        for index, values in enumerate(rows):
            joined = _row_text(values).lower()
            if not joined:
                continue
            for target, patterns in TARGET_PATTERNS.items():
                matched_pattern = next((pattern for pattern in patterns if pattern in joined), None)
                if matched_pattern is None:
                    continue

                row_number = index + 1
                previous_1 = rows[index - 1] if index >= 1 else tuple()
                previous_2 = rows[index - 2] if index >= 2 else tuple()
                matches.append(
                    {
                        "target": target,
                        "sheet": sheet.title,
                        "row": row_number,
                        "matched_pattern": matched_pattern,
                        "row_values": _row_cells(values, row_number),
                        "header_minus_1": _row_cells(previous_1, row_number - 1),
                        "header_minus_2": _row_cells(previous_2, row_number - 2),
                    }
                )
                break

    workbook.close()
    return matches
